"""
Excel report generation module.

Compiles all analysis results into a multi-sheet publication-ready Excel
workbook with formatting: Demographics, Knowledge, Likert Analysis,
Coping Strategies, Chi-Square Tests, and Summary.
"""

import io
import logging
from datetime import datetime
from typing import Any, Dict

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Styling constants
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=13, color="2B579A")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="404040")
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header_row(ws, row: int, max_col: int):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _style_data_rows(ws, start_row: int, end_row: int, max_col: int):
    """Apply body styling to data rows."""
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if col > 1 else LEFT_ALIGN


def _auto_column_width(ws, max_col: int, max_row: int):
    """Auto-fit column widths based on content."""
    for col in range(1, max_col + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 4, 55)
        ws.column_dimensions[col_letter].width = adjusted_width


def generate_report(
    metadata: Dict[str, Any],
    demographics: Dict[str, Any],
    knowledge: Dict[str, Any],
    likert: Dict[str, Any],
    coping: Dict[str, Any],
    inferential: Dict[str, Any],
) -> bytes:
    """
    Generate a multi-sheet Excel report.

    Returns:
        Bytes of the Excel file.
    """
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # ── Sheet 1: Demographics ────────────────────────────────────────
        _write_demographics_sheet(writer, demographics)

        # ── Sheet 2: Knowledge ───────────────────────────────────────────
        _write_knowledge_sheet(writer, knowledge)

        # ── Sheet 3: Likert Analysis ─────────────────────────────────────
        _write_likert_sheet(writer, likert)

        # ── Sheet 4: Coping Strategies ───────────────────────────────────
        _write_coping_sheet(writer, coping)

        # ── Sheet 5: Chi-Square Tests ────────────────────────────────────
        _write_inferential_sheet(writer, inferential)

        # ── Sheet 6: Summary ─────────────────────────────────────────────
        _write_summary_sheet(writer, metadata, likert)

    buffer.seek(0)
    return buffer.read()


def _write_demographics_sheet(writer, demographics: Dict):
    """Write demographic frequency tables."""
    rows = []
    demo_labels = {
        "sex": "Sex",
        "age_group": "Age Group",
        "marital_status": "Marital Status",
        "professional_cadre": "Professional Cadre (Strata)",
        "work_experience": "Years of Work Experience",
        "income_level": "Monthly Income / Grade Level",
    }

    for key, label in demo_labels.items():
        if key not in demographics:
            continue
        # Add variable header
        rows.append({"Variable": label, "Category": "", "Frequency (n)": "", "Percentage (%)": ""})
        for entry in demographics[key]:
            rows.append(
                {
                    "Variable": "",
                    "Category": entry["category"],
                    "Frequency (n)": entry["count"],
                    "Percentage (%)": entry["percentage"],
                }
            )
        # Add subtotal row
        total_count = sum(e["count"] for e in demographics[key])
        rows.append(
            {
                "Variable": "",
                "Category": "Total",
                "Frequency (n)": total_count,
                "Percentage (%)": 100.0,
            }
        )
        rows.append({"Variable": "", "Category": "", "Frequency (n)": "", "Percentage (%)": ""})

    df = pd.DataFrame(rows)
    sheet_name = "Demographics"
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value="Table 1: Socio-Demographic Characteristics of Respondents").font = TITLE_FONT
    _style_header_row(ws, 2, 4)
    _style_data_rows(ws, 3, 3 + len(rows), 4)
    _auto_column_width(ws, 4, 2 + len(rows))


def _write_knowledge_sheet(writer, knowledge: Dict):
    """Write knowledge/awareness frequency tables."""
    rows = []
    know_labels = {
        "awareness": "Item 7: Awareness of Fuel Subsidy Removal",
        "understanding": "Item 8: Understanding of Subsidy–Hospital Cost Relationship",
        "financial_burden": "Item 9: Belief in Increased Patient Financial Burden",
    }

    for key, label in know_labels.items():
        if key not in knowledge:
            continue
        rows.append({"Question": label, "Response": "", "Frequency (n)": "", "Percentage (%)": ""})
        for entry in knowledge[key]:
            rows.append(
                {
                    "Question": "",
                    "Response": entry["category"],
                    "Frequency (n)": entry["count"],
                    "Percentage (%)": entry["percentage"],
                }
            )
        total_count = sum(e["count"] for e in knowledge[key])
        rows.append({"Question": "", "Response": "Total", "Frequency (n)": total_count, "Percentage (%)": 100.0})
        rows.append({"Question": "", "Response": "", "Frequency (n)": "", "Percentage (%)": ""})

    df = pd.DataFrame(rows)
    sheet_name = "Knowledge"
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value="Table 2: Knowledge and Awareness of Fuel Subsidy Removal").font = TITLE_FONT
    _style_header_row(ws, 2, 4)
    _style_data_rows(ws, 3, 3 + len(rows), 4)
    _auto_column_width(ws, 4, 2 + len(rows))


def _write_likert_sheet(writer, likert: Dict):
    """Write Likert analysis table."""
    rows = []
    current_section = None

    for item in likert["items"]:
        if item["section"] != current_section:
            current_section = item["section"]
            rows.append(
                {
                    "Item": f"--- {current_section} ---",
                    "SA": "",
                    "A": "",
                    "D": "",
                    "SD": "",
                    "Weighted Mean": "",
                    "Decision": "",
                }
            )

        rows.append(
            {
                "Item": item["statement"],
                "SA": item["frequencies"]["SA"],
                "A": item["frequencies"]["A"],
                "D": item["frequencies"]["D"],
                "SD": item["frequencies"]["SD"],
                "Weighted Mean": item["weighted_mean"],
                "Decision": item["decision"],
            }
        )

    # Grand mean row
    rows.append({"Item": "", "SA": "", "A": "", "D": "", "SD": "", "Weighted Mean": "", "Decision": ""})
    rows.append(
        {
            "Item": "Overall Grand Mean",
            "SA": "",
            "A": "",
            "D": "",
            "SD": "",
            "Weighted Mean": likert["overall_mean"],
            "Decision": "Positive Effect" if likert["overall_mean"] >= 2.5 else "Negative Effect",
        }
    )

    # Section means
    for section_name, section_mean in likert.get("section_means", {}).items():
        rows.append(
            {
                "Item": f"Sub-Section Mean: {section_name}",
                "SA": "",
                "A": "",
                "D": "",
                "SD": "",
                "Weighted Mean": section_mean,
                "Decision": "Positive Effect" if section_mean >= 2.5 else "Negative Effect",
            }
        )

    df = pd.DataFrame(rows)
    sheet_name = "Likert Analysis"
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value="Table 3: Weighted Mean Analysis of Perceived Effects (Section C)").font = TITLE_FONT
    _style_header_row(ws, 2, 7)
    _style_data_rows(ws, 3, 3 + len(rows), 7)
    _auto_column_width(ws, 7, 2 + len(rows))


def _write_coping_sheet(writer, coping: Dict):
    """Write coping strategies frequency tables."""
    rows = []
    coping_labels = {
        "personal_transport": "Item 16: Personal Transportation Adjustments",
        "institutional": "Item 17: Institutional Adjustments",
        "financial": "Item 18: Financial Coping Strategies",
    }

    for key, label in coping_labels.items():
        if key not in coping:
            continue
        rows.append({"Category": label, "Strategy": "", "Frequency (n)": "", "Percentage (%)": ""})
        for entry in coping[key]:
            rows.append(
                {
                    "Category": "",
                    "Strategy": entry["strategy"],
                    "Frequency (n)": entry["count"],
                    "Percentage (%)": entry["percentage"],
                }
            )
        rows.append({"Category": "", "Strategy": "", "Frequency (n)": "", "Percentage (%)": ""})

    df = pd.DataFrame(rows)
    sheet_name = "Coping Strategies"
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value="Table 4: Coping Strategies Adopted by Healthcare Workers (Section D)").font = TITLE_FONT
    _style_header_row(ws, 2, 4)
    _style_data_rows(ws, 3, 3 + len(rows), 4)
    _auto_column_width(ws, 4, 2 + len(rows))


def _write_inferential_sheet(writer, inferential: Dict):
    """Write chi-square / Fisher's Exact test results."""
    tests = inferential.get("chi_square_tests", [])
    rows = []
    for test in tests:
        rows.append(
            {
                "Test": test["test_name"],
                "Variables": test["variables"],
                "χ² Statistic": test["chi2"] if test["chi2"] is not None else "N/A",
                "df": test["dof"] if test["dof"] is not None else "N/A",
                "p-value": test["p_value"] if test["p_value"] is not None else "N/A",
                "Significant (α=0.05)": (
                    "Yes" if test["significant"] else "No"
                )
                if test["significant"] is not None
                else "N/A",
                "Method": test["method"],
                "Notes": test["note"] or "",
            }
        )

    df = pd.DataFrame(rows)
    sheet_name = "Chi-Square Tests"
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value="Table 5: Chi-Square Tests of Independence (Professional Cadre)").font = TITLE_FONT
    _style_header_row(ws, 2, 8)
    _style_data_rows(ws, 3, 3 + len(rows), 8)
    _auto_column_width(ws, 8, 2 + len(rows))


def _write_summary_sheet(writer, metadata: Dict, likert: Dict):
    """Write study summary/metadata sheet."""
    summary_data = [
        ("Study Title", "Perceived Effects of Fuel Subsidy Removal on Healthcare Delivery Among Health Workers in Prince Audu Abubakar University Teaching Hospital, Anyigba"),
        ("Date of Analysis", datetime.now().strftime("%B %d, %Y")),
        ("", ""),
        ("RESPONSE METRICS", ""),
        ("Total Sample Size (Distributed)", metadata.get("total_sample_size", "")),
        ("Completed Responses", metadata.get("completed_responses", "")),
        ("Non-Respondents (No Consent)", metadata.get("non_respondents", "")),
        ("Response Rate", f"{metadata.get('response_rate', '')}%"),
        ("", ""),
        ("METHODOLOGY", ""),
        ("Research Design", "Descriptive Cross-Sectional"),
        ("Sampling Technique", "Stratified Random Sampling"),
        ("Instrument", "Self-structured questionnaire (4 sections: A–D)"),
        ("Likert Scale", "4-point: Strongly Agree (4), Agree (3), Disagree (2), Strongly Disagree (1)"),
        ("Decision Threshold", "Mean ≥ 2.5 = Positive Effect | Mean < 2.5 = Negative Effect"),
        ("Imputation Method", "Proportional Stochastic Imputation (seed=42)"),
        ("Inferential Tests", "Chi-Square Test of Independence with Fisher's Exact fallback (α = 0.05)"),
        ("Likert Binarization", "SA + A → Impacted | D + SD → Not Impacted (for chi-square)"),
        ("", ""),
        ("KEY FINDINGS", ""),
        ("Overall Grand Mean (Likert)", likert.get("overall_mean", "")),
    ]

    # Add section means
    for section_name, section_mean in likert.get("section_means", {}).items():
        summary_data.append((f"  Sub-Section Mean: {section_name}", section_mean))

    df = pd.DataFrame(summary_data, columns=["Field", "Value"])
    sheet_name = "Summary"
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

    ws = writer.sheets[sheet_name]
    # Style
    for row in range(1, len(summary_data) + 2):
        for col in range(1, 3):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.alignment = LEFT_ALIGN
            # Bold section headers
            if cell.value and str(cell.value).isupper() and col == 1:
                cell.font = SUBTITLE_FONT

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 80

    logger.info("Report generated successfully with 6 sheets.")
